from django.core.management.base import BaseCommand
from django.utils.dateparse import parse_datetime, parse_date
from openpyxl import load_workbook
from datetime import datetime

from order.models import Order, OrderItem
from item.models import Item
from aircraft.models import Aircraft
from django.contrib.auth.models import User


class Command(BaseCommand):
    help = "Importa planilha Pedidos Contrato 89 (FSM) para Order e OrderItem"

    def handle(self, *args, **options):
        path = "static/spreadsheet/pedidos_FSM_contrato_89.xlsx"
        username = "admin"

        # Buscar usuário
        try:
            user = User.objects.get(username=username)
        except User.DoesNotExist:
            self.stderr.write(self.style.ERROR(f"Usuário {username} não encontrado"))
            return

        # Carregar planilha
        wb = load_workbook(path, data_only=True)
        ws = wb["Pedidos Contrato 89"]

        # Índices das colunas
        COL_PEDIDO = 1
        COL_DATA_PEDIDO = 2
        COL_SOLICITANTE = 3
        COL_OPERADOR = 4
        COL_ANV = 5
        COL_TIPO_ATD = 6
        COL_TIPO_PED = 7
        COL_MPN = 8
        COL_NOME = 9
        COL_QTD = 10
        COL_DOC_TS = 11
        COL_MOTIVO = 12
        COL_OBS_ITEM = 13
        COL_TROUBLESHOOTING = 14
        COL_TSN = 15
        COL_TSO = 16
        COL_SN = 17
        COL_ANV_TSN = 18
        COL_VENC = 19
        COL_DEST_ANV = 20
        COL_PN_ALT1 = 21
        COL_PN_ALT2 = 22
        COL_CONTRATO_ANT = 23
        COL_DPE = 24
        COL_STATUS = 25
        COL_SN_RECEBIDO = 26
        COL_NF_TRANSP = 27
        COL_DATA_ATD = 28
        COL_QTD_FORN = 29
        COL_QTD_FALT = 30
        COL_OBS1 = 31

        # Contadores
        created_orders = 0
        updated_orders = 0
        created_items = 0

        # Conversões
        def to_date(v):
            if not v:
                return None

            # Caso seja datetime → converte para date
            if isinstance(v, datetime):
                return v.date()

            # Caso seja date
            if hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day"):
                return v

            # Strings
            if isinstance(v, str):
                v = v.strip()

                dt = parse_datetime(v)
                if dt:
                    return dt.date()

                dt = parse_date(v.split(" ")[0])
                if dt:
                    return dt

            return None


        def to_decimal(v):
            if v in (None, ""):
                return None
            try:
                return float(str(v).replace(",", "."))
            except:
                return None

        # Filtro de aeronave
        def aircraft_filter(s):
            mapa = {
                "5001": "5001",
                "5002": "5002",
                "5003": "5003",
                "5005": "5005",
                "5007": "5007",
                "5008": "5008",
                "5013": "5013",
            }

            s = s.upper()

            for k in mapa:
                if k in s or ("EB" + k) in s:
                    return Aircraft.objects.get_or_create(numeral=mapa[k])[0]

            if "KAN" in s:
                return Aircraft.objects.get_or_create(numeral="KAN")[0]

            return Aircraft.objects.get_or_create(numeral="5001")[0]

        # Loop principal

        self.stdout.write(f"Iniciando importação de {path}...")
        for row in range(2, ws.max_row + 1):

            num_pedido = ws.cell(row=row, column=COL_PEDIDO).value
            if not num_pedido:
                continue

            try:
                num_pedido = int(num_pedido)
            except:
                continue

            # Carregar linha
            data_pedido_raw = ws.cell(row=row, column=COL_DATA_PEDIDO).value
            solicitante_raw = ws.cell(row=row, column=COL_SOLICITANTE).value
            operador = ws.cell(row=row, column=COL_OPERADOR).value
            anv = ws.cell(row=row, column=COL_ANV).value
            tipo_atd_raw = ws.cell(row=row, column=COL_TIPO_ATD).value
            tipo_ped_raw = ws.cell(row=row, column=COL_TIPO_PED).value
            mpn = ws.cell(row=row, column=COL_MPN).value
            nome_item = ws.cell(row=row, column=COL_NOME).value
            qtd = ws.cell(row=row, column=COL_QTD).value
            doc_raw = ws.cell(row=row, column=COL_DOC_TS).value
            motivo = ws.cell(row=row, column=COL_MOTIVO).value
            obs_item = ws.cell(row=row, column=COL_OBS_ITEM).value
            troubleshooting = ws.cell(row=row, column=COL_TROUBLESHOOTING).value
            tsn = ws.cell(row=row, column=COL_TSN).value
            tso = ws.cell(row=row, column=COL_TSO).value
            sn_recebido = ws.cell(row=row, column=COL_SN_RECEBIDO).value
            venc_raw = ws.cell(row=row, column=COL_VENC).value
            dest_anv = ws.cell(row=row, column=COL_DEST_ANV).value
            contrato_ant_raw = ws.cell(row=row, column=COL_CONTRATO_ANT).value
            dpe = ws.cell(row=row, column=COL_DPE).value
            status_raw = ws.cell(row=row, column=COL_STATUS).value
            nf = ws.cell(row=row, column=COL_NF_TRANSP).value
            data_atd_raw = ws.cell(row=row, column=COL_DATA_ATD).value
            qtd_forn = ws.cell(row=row, column=COL_QTD_FORN).value
            obs1 = ws.cell(row=row, column=COL_OBS1).value

            # Conversões
            order_date = to_date(data_pedido_raw)
            attended_date = to_date(data_atd_raw)
            venc_date = to_date(venc_raw)

        

            if not order_date:
                continue

            # --------------------------
            # MAPEAMENTOS
            # --------------------------

            requester = None
            if solicitante_raw:
                txt = str(solicitante_raw).upper()
                if "1BAVEX" in txt or "BAVEX" in txt:
                    requester = "1BAVEX"
                elif "BMS" in txt:
                    requester = "BMS"

            order_type = None
            if tipo_ped_raw:
                t = str(tipo_ped_raw).upper()
                if "FSM" in t:
                    order_type = "FSM"
                elif "RMS" in t:
                    order_type = "RMS"
                elif "REQ" in t:
                    order_type = "REQ"

            status = "OPEN"
            if status_raw:
                t = str(status_raw).upper()
                if "ATENDIDO PARCIALMENTE" in t:
                    status = "OPEN2"
                elif "NÃO ATENDIDO" in t or "NAO ATENDIDO" in t:
                    status = "CLOSE2"
                elif "ATENDIDO" in t:
                    status = "CLOSE"

            # Aeronaves
            anv_final = aircraft_filter(str(anv)) if anv else None
            dest_anv_final = aircraft_filter(str(dest_anv)) if dest_anv else None

            # DOC/TEC PUB
            doc_final = None
            tec_pub_final = None
            if doc_raw:
                raw = str(doc_raw).strip()
                if raw != "-":
                    upper = raw.upper()
                    if "IPC" in upper:
                        doc_final = "IPC"
                        tec_pub_final = upper.replace("IPC", "").strip()
                    elif "ECMM" in upper:
                        doc_final = "ECMM"
                        tec_pub_final = upper.replace("ECMM", "").strip()
                    elif "MMA" in upper:
                        doc_final = "MMA"
                        tec_pub_final = upper.replace("MMA", "").strip()
                    else:
                        tec_pub_final = raw

            # --------------------------
            # CRIAR / ATUALIZAR ORDER
            # --------------------------

            order, created = Order.objects.get_or_create(
                order_number=num_pedido,
                order_year=order_date.year,
                defaults={
                    "order_date": order_date,
                    "requester": requester,
                    "order_type": order_type,
                    "status": status,
                    "notes": str(obs1) if obs1 else "",
                    "created_by": user,
                    "updated_by": user,
                },
            )

            if created:
                created_orders += 1
            else:
                changed = False
                if requester and not order.requester:
                    order.requester = requester
                    changed = True
                if order_type and not order.order_type:
                    order.order_type = order_type
                    changed = True
                if status and status != order.status:
                    order.status = status
                    changed = True
                if obs1:
                    order.notes = (order.notes or "") + "\n" + str(obs1)
                    changed = True
                if changed:
                    order.updated_by = user
                    order.save()
                    updated_orders += 1

            # --------------------------
            # ITEM
            # --------------------------
            item_obj = None

            if mpn:
                mpn_str = str(mpn).strip()

                item_obj, created_item = Item.objects.get_or_create(
                    mpn=mpn_str,
                    defaults={
                        "name": str(nome_item).strip() if nome_item else mpn_str,
                        "doc": doc_final,
                        "tec_pub": tec_pub_final,
                        "created_by": user,
                    },
                )

                # Atualizar doc/tec_pub
                changed = False
                if doc_final and item_obj.doc != doc_final:
                    item_obj.doc = doc_final
                    changed = True
                if tec_pub_final and item_obj.tec_pub != tec_pub_final:
                    item_obj.tec_pub = tec_pub_final
                    changed = True
                if changed:
                    item_obj.save()


            # --------------------------
            # Campos adicionais
            # --------------------------

            contract_old = False
            if contrato_ant_raw:
                t = str(contrato_ant_raw).strip().lower()
                contract_old = t.startswith("s") or t == "1" or "yes" in t or "sim" in t

            tsn_val = to_decimal(tsn)
            tso_val = to_decimal(tso)

            try:
                quantity = int(qtd) if qtd else 1
            except:
                quantity = 1

            # --------------------------
            # CRIAR ORDERITEM
            # --------------------------

            try:
                OrderItem.objects.create(
                    order=order,
                    item_item=item_obj,
                    operator=str(operador).strip() if operador else None,
                    aircraft=anv_final,
                    aircraft_destination=dest_anv_final,
                    service_type=str(tipo_atd_raw).upper().strip() if tipo_atd_raw else None,
                    quantity=quantity,
                    quantity_supplied=qtd_forn,
                    dpe=str(dpe).strip() if dpe else None,
                    nf_answer=nf,
                    sn_attended=sn_recebido,
                    expiration_date_attended=venc_date,
                    attended_date=attended_date,
                    contract_old=contract_old,
                    reason=str(motivo).strip() if motivo else None,
                    troubleshooting=str(troubleshooting).strip() if troubleshooting else None,
                    observation=str(obs_item).strip() if obs_item else None,
                    created_by=user,
                    tsn_item=tsn_val,
                    tso_item=tso_val,
                )
                created_items += 1
            except Exception as e:
                self.stderr.write(f"Erro ao criar OrderItem na linha {row}: {str(e)}")

        # Final
        self.stdout.write(
            self.style.SUCCESS(
                f"Importação concluída!\n"
                f"Pedidos criados: {created_orders}\n"
                f"Pedidos atualizados: {updated_orders}\n"
                f"Itens criados: {created_items}"
            )
        )
