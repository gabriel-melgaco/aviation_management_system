from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.views import View
from django.db.models import Q
from . import models
from . import forms
from django.urls import reverse_lazy
from django.utils import timezone
from django.shortcuts import get_object_or_404
from django.contrib import messages
from django.shortcuts import redirect
from openpyxl import load_workbook
from django.http import FileResponse
from io import BytesIO





class OrderListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = models.Order
    template_name = 'order_list.html'
    context_object_name = 'orders'
    paginate_by = 10
    permission_required = 'order.view_order'

    def get_queryset(self):
        queryset = models.Order.objects.all().prefetch_related(
            'items',  # OrderItems relacionados
            'items__inventory_item',  # Item de inventário
            'items__inventory_item__item',  # Item dentro do inventário
            'items__item_item',  # Item FMS direto
        ).select_related(
            'created_by',
            'updated_by'
        ).order_by('-order_date')

        # Filtro de busca por texto
        search = self.request.GET.get('search', '').strip()
        if search:
            queryset = queryset.filter(
                Q(order_number__icontains=search) |
                Q(order_year__icontains=search) |
                Q(requester__icontains=search) |
                Q(notes__icontains=search) |
                # Busca nos OrderItems - Item de Inventário
                Q(items__inventory_item__item__name__icontains=search) |
                Q(items__inventory_item__item__mpn__icontains=search) |
                Q(items__inventory_item__item__pn__icontains=search) |
                # Busca nos OrderItems - Item FMS
                Q(items__item_item__name__icontains=search) |
                Q(items__item_item__mpn__icontains=search) |
                Q(items__item_item__pn__icontains=search) |
                # Busca por outros campos de OrderItem
                Q(items__dpe__icontains=search) |
                Q(items__eglog__icontains=search) |
                Q(items__log__icontains=search) |
                Q(items__operator__icontains=search)
            ).distinct().order_by('-order_date')

        # Filtro por status
        status = self.request.GET.get('status', '').strip()
        if status:
            queryset = queryset.filter(status=status).order_by('-order_date')

        # Filtro por tipo de pedido
        order_type = self.request.GET.get('order_type', '').strip()
        if order_type:
            queryset = queryset.filter(order_type=order_type).order_by('-order_date')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Estatísticas gerais (sem filtros aplicados)
        all_orders = models.Order.objects.all()
        
        context['total_orders'] = all_orders.count()
        context['open_orders'] = all_orders.filter(status='OPEN').count()
        context['closed_orders'] = all_orders.filter(status='CLOSE').count()
        context['rms_orders'] = all_orders.filter(order_type='RMS').count()
        context['fsm_orders'] = all_orders.filter(order_type='FSM').count()

        
        # Adicionar filtros ativos
        context['active_filters'] = {
            'search': self.request.GET.get('search', ''),
            'status': self.request.GET.get('status', ''),
            'order_type': self.request.GET.get('order_type', ''),
        }
        
        return context


class OrderCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = models.Order
    template_name = 'order_create.html'
    form_class = forms.OrderForm
    success_url = reverse_lazy('order_list')
    permission_required = 'order.add_order'


    def order_number_creator(self):
        current_year = timezone.now().year

        last_order = models.Order.objects.filter(order_year=current_year).order_by('-order_number').first()

        if last_order:
            return last_order.order_number + 1
        else:
            return 1

    def form_valid(self, form):
        current_datetime = timezone.now()
        year = current_datetime.year
        form.instance.order_year = year
        if self.request.user.is_authenticated:
            form.instance.created_by = self.request.user
        form.instance.order_number = self.order_number_creator()
        return super().form_valid(form)


class OrderUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = models.Order
    template_name = 'order_update.html'
    form_class = forms.OrderForm
    success_url = reverse_lazy('order_list')
    permission_required = 'order.change_order'


class OrderDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = models.Order
    template_name = 'order_delete.html'
    success_url = reverse_lazy('order_list')
    permission_required = 'order.delete_order'


class OrderDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = models.Order
    template_name = 'order_detail.html'
    context_object_name = 'object'
    permission_required = 'order.view_order'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Adicionar os itens do pedido ao contexto
        context['order_items'] = models.OrderItem.objects.filter(
            order=self.object
        ).select_related(
            'inventory_item',
            'inventory_item__item',
            'item_item',
            'aircraft',
            'created_by'
        ).order_by('-created_at')
        
        # Estatísticas dos itens
        context['total_items'] = context['order_items'].count()
        context['total_quantity'] = sum(item.quantity for item in context['order_items'])
        context['collected_items'] = context['order_items'].filter(collected=True).count()
        
        # Itens por tipo de serviço
        context['rush_items'] = context['order_items'].filter(service_type='RUSH').count()
        context['aog_items'] = context['order_items'].filter(service_type='AOG').count()
        context['prog_items'] = context['order_items'].filter(service_type='PROG').count()
        
        return context



# -----------------------Order Item Views------------------------------------
class OrderItemListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = models.OrderItem
    template_name = 'order_detail.html'
    context_object_name = 'orders_itens'
    permission_required = 'order.view_orderitem'


class OrderItemCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = models.OrderItem
    form_class = forms.OrderItemForm
    template_name = 'order_item_create.html'
    permission_required = 'order.add_orderitem'

    
    def dispatch(self, request, *args, **kwargs):
        # Buscar o pedido (Order) pela pk passada na URL
        self.order = get_object_or_404(models.Order, pk=kwargs.get('pk'))
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['order'] = self.order
        return context
    
    def form_valid(self, form):
        form.instance.order = self.order
        if self.request.user.is_authenticated:
            form.instance.created_by = self.request.user
                       
        return super().form_valid(form)
    
    
    def get_success_url(self):
        messages.success(
        self.request,
        f'Item adicionado ao pedido {self.order.order_number}/{self.order.order_year} com sucesso!'
    )       
        return reverse_lazy('order_detail', kwargs={'pk': self.order.pk})
        


class OrderItemUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = models.OrderItem
    template_name = 'order_item_create.html'
    form_class = forms.OrderItemForm
    permission_required = 'order.change_orderitem'


    def dispatch(self, request, *args, **kwargs):
        # Buscar o item de pedido (OrderItem) e o pedido (Order) relacionado
        self.item = get_object_or_404(models.OrderItem, pk=kwargs.get('pk'))
        self.order = self.item.order
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['order'] = self.order
        return context
    
    def get_success_url(self):
        messages.success(
            self.request,
            f'Item atualizado no pedido {self.order.order_number}/{self.order.order_year} com sucesso!'
        )       
        return reverse_lazy('order_detail', kwargs={'pk': self.order.pk})



class OrderItemDeleteView(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'order.delete_orderitem'

    def post(self, request, pk, *args, **kwargs):
        item = get_object_or_404(models.OrderItem, pk=pk)
        
        order_id = item.order.pk
        
        item.delete()
        
        return redirect('order_detail', pk=order_id)
    


class OrderExportArchive(LoginRequiredMixin, PermissionRequiredMixin, View):
    permission_required = 'order.view_order'

    def get(self, request, pk):
        # Acessando os dados no banco
        order = get_object_or_404(models.Order, pk=pk)
        order_items = models.OrderItem.objects.filter(order=order).select_related(
            'aircraft', 
            'inventory_item__item',
            'item_item'
        )

        # Caminho da planilha template
        path_archive = "static/spreadsheet/order_spreadsheet.xlsx"

        # Carregar a planilha existente
        wb = load_workbook(path_archive)
        ws = wb.active
        
        # Começar na linha 2 (assumindo que linha 1 é cabeçalho)
        linha = 2

        # Listagem dos itens
        for item in order_items:
            # Coluna 1: Solicitante
            ws.cell(row=linha, column=1, value='1BAV')
            
            # Coluna 2: Aeronave (verificar se existe)
            ws.cell(row=linha, column=2, value=item.aircraft.numeral if item.aircraft else '')
            
            # Coluna 3: Tipo de Atendimento
            ws.cell(row=linha, column=3, value=item.service_type if item.service_type else '')
            
            # Coluna 4: Tipo de Pedido
            ws.cell(row=linha, column=4, value=order.order_type if order.order_type else '')
            
            # Colunas 5 e 6: MPN e Nome do Item
            # Verificar se é item de inventário ou item direto
            if item.inventory_item and item.inventory_item.item:
                mpn = item.inventory_item.item.mpn
                name = item.inventory_item.item.name
            elif item.item_item:
                mpn = item.item_item.mpn
                name = item.item_item.name
            else:
                mpn = ''
                name = ''
            
            ws.cell(row=linha, column=5, value=mpn)
            ws.cell(row=linha, column=6, value=name)

            # Coluna 7: Quantidade
            ws.cell(row=linha, column=7, value=item.quantity if item.quantity else '')
            
            # Coluna 8: DOC/PUB TEC
            ws.cell(row=linha, column=8, value=f"{item.item_item.doc} {item.item_item.tec_pub}" if item.item_item else f"{item.inventory_item.item.doc} {item.inventory_item.item.tec_pub}" if item.inventory_item else '')
            
            # Coluna 9: Motivo
            ws.cell(row=linha, column=9, value=item.reason if item.reason else '')

            # Coluna 10: Observação
            ws.cell(row=linha, column=10, value=item.observation if item.observation else '')

             # Coluna 11: Descrição Pane
            ws.cell(row=linha, column=11, value=item.failure_description if item.failure_description else '')

            # Coluna 12: Troubleshooting
            ws.cell(row=linha, column=12, value=item.troubleshooting if item.troubleshooting else '')
            
            # Coluna 13: TSN
            ws.cell(row=linha, column=13, value=item.tsn_item if item.tsn_item else '')

            # Coluna 14: TS0
            ws.cell(row=linha, column=14, value=item.tso_item if item.tso_item else '')

            # Coluna 15: Serial Number, SFC
            ws.cell(row=linha, column=15, value=item.inventory_item.serial_number if item.inventory_item else '')
            
            # Coluna 16: TSN Anv
            ws.cell(row=linha, column=16, value=item.aircraft.tsn if item.aircraft and item.aircraft.tsn else '')

            # Coluna 17: Vencimento, SFC
            ws.cell(row=linha, column=17, value=item.inventory_item.expiration_date.date() if item.inventory_item else '')
            
            # Coluna 18: Anv de destino
            ws.cell(row=linha, column=18, value=item.aircraft_destination.numeral if item.aircraft_destination else '')

            # Coluna 19 e 20: PN Alternativo
            item_alt = None
            if item.inventory_item:
                item_alt = item.inventory_item.item
            elif item.item_item:
                item_alt = item.item_item

            if item_alt:
                equivalent = list(item_alt.all_equivalents)

            ws.cell(row=linha, column=19, value=equivalent[0].mpn if len(equivalent) > 0  else '')
            ws.cell(row=linha, column=20, value=equivalent[1].mpn if len(equivalent) > 1  else '')

            linha += 1

        # Salvar na memória
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Retornar o download
        response = FileResponse(
            buffer,
            as_attachment=True,
            filename=f"Pedido_{order.order_year}-{order.order_number}.xlsx"
        )
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        return response

    
