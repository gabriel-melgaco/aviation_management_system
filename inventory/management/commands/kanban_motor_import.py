from django.core.management.base import BaseCommand
from django.utils.dateparse import parse_datetime, parse_date
from openpyxl import load_workbook
from datetime import datetime

from inventory.models import Inventory
from item.models import Item
from location.models import LocationSite, Location
from aircraft.models import Aircraft
from django.contrib.auth.models import User


class Command(BaseCommand):
    help = "Importa Kanban Motor"

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='static/spreadsheet/kanban_motor.xlsx',
            help='Caminho do arquivo Excel'
        )
        parser.add_argument(
            '--sheet',
            type=str,
            default='KANBAN_MOTOR',
            help='Nome da aba da planilha'
        )
        parser.add_argument(
            '--user',
            type=str,
            default='admin',
            help='Username do usuário para created_by'
        )

    def handle(self, *args, **options):
        path = options['file']
        sheet_name = options['sheet']
        username = options['user']

        # Usuário
        try:
            user = User.objects.get(username=username)
        except User.DoesNotExist:
            self.stderr.write(self.style.ERROR(f"Usuário {username} não encontrado"))
            return

        # Planilha
        try:
            wb = load_workbook(path, data_only=True)
            ws = wb[sheet_name]
        except FileNotFoundError:
            self.stderr.write(self.style.ERROR(f"Arquivo {path} não encontrado"))
            return
        except KeyError:
            self.stderr.write(self.style.ERROR(f"Aba {sheet_name} não encontrada"))
            return

        COL_CASE = 1
        COL_ITEM_LOC = 2
        COL_MPN = 3
        COL_NOME = 4
        COL_QTD = 5
        COL_CAPITULO = 6
        COL_FIGURA = 7
        COL_ITEM = 8
        COL_QTD_MIN = 9

        created_items = 0
        created_inventory = 0
        errors = 0

        def to_int(v):
            if v in (None, "", "-"):
                return None
            try:
                return int(float(str(v).replace(",", ".")))
            except:
                return None

        def clean_text(v):
            if v in (None, "", "-"):
                return None
            return str(v).strip()

        self.stdout.write(f"Iniciando importação de {path}...")

        for row in range(2, ws.max_row + 1):
            try:
                case = clean_text(ws.cell(row=row, column=COL_CASE).value)
                item_loc = clean_text(ws.cell(row=row, column=COL_ITEM_LOC).value)
                mpn = clean_text(ws.cell(row=row, column=COL_MPN).value)
                nome = clean_text(ws.cell(row=row, column=COL_NOME).value)
                qtd = to_int(ws.cell(row=row, column=COL_QTD).value) or 1
                qtd_min = to_int(ws.cell(row=row, column=COL_QTD_MIN).value)

                capitulo = clean_text(ws.cell(row=row, column=COL_CAPITULO).value)
                figura = clean_text(ws.cell(row=row, column=COL_FIGURA).value)
                item_ref = clean_text(ws.cell(row=row, column=COL_ITEM).value)

                if not mpn:
                    continue
                
                if not nome:
                    nome = f"{mpn}"

                if not capitulo or not figura or not item_ref:
                    doc = None
                    tec_pub = None
                else:
                    doc = "IETP"
                    tec_pub = f"{capitulo}-{figura}-{item_ref}"

                item_obj, created = Item.objects.get_or_create(
                    mpn=mpn,
                    defaults={
                        "name": nome,
                        "doc": doc,
                        "tec_pub": tec_pub,
                        "created_by": user,
                    }
                )

                if created:
                    created_items += 1
                else:
                    updated = False
                    if item_obj.doc != doc:
                        item_obj.doc = doc
                        updated = True
                    if item_obj.tec_pub != tec_pub:
                        item_obj.tec_pub = tec_pub
                        updated = True
                    if updated:
                        item_obj.save()

                site, _ = LocationSite.objects.get_or_create(
                    location_site="1bavex",
                    location_sub_site="spu",
                    defaults={"type": "internal"},
                )

                location, _ = Location.objects.get_or_create(
                    om=site,
                    section="KANBAN MOTOR",
                    shelf=None,
                    case=case,
                    item_number=item_loc,
                )

                Inventory.objects.create( 
                    item=item_obj, 
                    serial_number=None,
                    kanban="ENGINE", 
                    location=location, 
                    quantity=qtd, 
                    minimum_quantity=qtd_min, 
                    expiration_date=None, 
                )

                created_inventory += 1

                if row % 100 == 0:
                    self.stdout.write(f"{row} linhas processadas...")

            except Exception as e:
                errors += 1
                self.stderr.write(
                    self.style.ERROR(f"Linha {row}: {str(e)}")
                )

        self.stdout.write(
            self.style.SUCCESS(
                f"\n{'='*60}\n"
                f"IMPORTAÇÃO FINALIZADA\n"
                f"{'='*60}\n"
                f"📦 Itens criados: {created_items}\n"
                f"📋 Registros de inventário: {created_inventory}\n"
                f"❌ Erros: {errors}\n"
                f"{'='*60}"
            )
        )
