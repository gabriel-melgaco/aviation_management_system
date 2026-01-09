from django.core.management.base import BaseCommand
from django.utils.dateparse import parse_datetime, parse_date
from openpyxl import load_workbook
from datetime import datetime

from order.models import Order, OrderItem
from item.models import Item
from aircraft.models import Aircraft
from django.contrib.auth.models import User


class Command(BaseCommand):
    help = "Importa planilha SPU"

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='static/spreadsheet/spu.xlsx',
            help='Caminho do arquivo Excel'
        )
        parser.add_argument(
            '--sheet',
            type=str,
            default='PRATELEIRAS',
            help='Nome da aba da planilha'
        )
        parser.add_argument(
            '--user',
            type=str,
            default='admin',
            help='Username do usuário para created_by'
        )

    def handle(self, *args, **options):
        path = options['file']
        sheet_name = options['sheet']
        username = options['user']

        # Buscar usuário
        try:
            user = User.objects.get(username='admin')
        except User.DoesNotExist:
            self.stderr.write(self.style.ERROR(f"Usuário {username} não encontrado"))
            return

        # Carregar planilha
        try:
            wb = load_workbook(path, data_only=True)
            ws = wb[sheet_name]
        except FileNotFoundError:
            self.stderr.write(self.style.ERROR(f"Arquivo {path} não encontrado"))
            return
        except KeyError:
            self.stderr.write(self.style.ERROR(f"Aba {sheet_name} não encontrada"))
            return

        COL_PEDIDO = 1          # PEDIDOS
        COL_DATA_PEDIDO = 2     # Data Pedido
        COL_SOLICITANTE = 3     # solicitante
        COL_OPERADOR = 4        # operador
        COL_ANV = 5             # ANV
        COL_TIPO_ATD = 6        # Tipo Atd
        COL_TIPO_PED = 7        # Tipo Ped
        COL_MPN = 8             # MPN
        COL_NOME = 9            # NOME
        COL_QTD = 10            # QTD
        COL_DOC_TS = 11         # DOC TSN-TSO
        COL_MOTIVO = 12         # MOTIVO
        COL_OBS = 13            # obs
        COL_DESC_PANE = 14      # DESC PANE
        COL_TROUBLESHOOTING = 15# TROUBLESHOOTING
        COL_TSN = 16            # TSN - Item
        COL_TSO = 17            # TSO - Item
        COL_SN = 18             # SN - Item
        COL_ANV_TSN = 19        # TSN - Anv
        COL_VENC = 20           #Data Venc
        COL_DEST_ANV = 21       # DEST ANV
        COL_PN_ALT1 = 22        # PN ALT 1
        COL_PN_ALT2 = 23        # PN ALT 2
        COL_DPE = 24            # DPE
        COL_LOG = 25            # LOG CARD
        COL_GMM = 26            #GMM
        COL_COLLECTED = 27       #COLETADO?
        COL_STATUS = 28         # STATUS
        COL_NF = 29             # NF
        COL_DATA_ATD = 30       # DATA ATD
        COL_CONTRATO_ANT = 31   #CONTRATO ANTERIOR
        COL_NOTES = 32          # ANOTAÇÕES

        # Contadores
        created_orders = 0
        updated_orders = 0
        created_items = 0
        errors = 0

        # Funções auxiliares
        def to_date(v):
            """Converte diversos formatos para date"""
            if not v:
                return None

            # Caso seja datetime
            if isinstance(v, datetime):
                return v.date()

            # Caso seja date
            if hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day"):
                return v

            # Strings
            if isinstance(v, str):
                v = v.strip()
                if not v or v == "-":
                    return None

                # Tenta parse datetime
                dt = parse_datetime(v)
                if dt:
                    return dt.date()

                # Tenta parse date
                dt = parse_date(v.split(" ")[0])
                if dt:
                    return dt

            return None

        def to_decimal(v):
            """Converte para decimal/float"""
            if v in (None, "", "-"):
                return None
            try:
                return float(str(v).replace(",", "."))
            except:
                return None

        def to_int(v):
            """Converte para inteiro"""
            if v in (None, "", "-"):
                return None
            try:
                return int(float(str(v).replace(",", ".")))
            except:
                return None

        def clean_text(v):
            """Limpa e retorna texto"""
            if v in (None, "", "-"):
                return None
            return str(v).strip()

        def aircraft_filter(s):
            """Identifica aeronave baseado no texto"""
            if not s:
                return None

            s = str(s).upper().strip()
            
            # Mapeamento de aeronaves
            aircraft_map = {
                "5001": "5001",
                "5002": "5002",
                "5003": "5003",
                "5005": "5005",
                "5007": "5007",
                "5008": "5008",
                "5013": "5013",
            }

            # Busca por número
            for key, numeral in aircraft_map.items():
                if key in s or f"EB{key}" in s:
                    return Aircraft.objects.get_or_create(numeral=numeral)[0]

            # Casos especiais
            if "KAN" in s or "KANARIS" in s:
                return Aircraft.objects.get_or_create(numeral="KAN")[0]

            # Padrão
            return Aircraft.objects.get_or_create(numeral="5001")[0]

        def parse_requester(v):
            """Identifica solicitante"""
            if not v:
                return None
            txt = str(v).upper()
            if "1BAVEX" in txt or "BAVEX" in txt or "1º BAVEX" in txt:
                return "1BAVEX"
            elif "BMS" in txt:
                return "BMS"
            return None

        def parse_order_type(v):
            """Identifica tipo de pedido"""
            if not v:
                return None
            t = str(v).upper()
            if "FSM" in t:
                return "FSM"
            elif "RMS" in t:
                return "RMS"
            elif "REQ" in t:
                return "REQ"
            return None

        def parse_status(v):
            """Identifica status do pedido"""
            if not v:
                return "OPEN"
            
            t = str(v).upper()
            if "ATENDIDO PARCIALMENTE" in t or "PARCIAL" in t:
                return "OPEN2"
            elif "NÃO ATENDIDO" in t or "NAO ATENDIDO" in t:
                return "CLOSE2"
            elif "ATENDIDO" in t:
                return "CLOSE"
            return "OPEN"

        def parse_doc_tec_pub(v):
            """Extrai DOC e TEC_PUB do campo"""
            if not v:
                return None, None
            
            raw = str(v).strip()
            if raw == "-" or not raw:
                return None, None
            
            upper = raw.upper()
            
            # Identifica tipo de documento
            if "IPC" in upper:
                return "IPC", upper.replace("IPC", "").strip()
            elif "ECMM" in upper:
                return "ECMM", upper.replace("ECMM", "").strip()
            elif "MMA" in upper:
                return "MMA", upper.replace("MMA", "").strip()
            elif "AMM" in upper:
                return "AMM", upper.replace("AMM", "").strip()
            elif "CMM" in upper:
                return "CMM", upper.replace("CMM", "").strip()
            else:
                return None, raw

        def parse_boolean(v):
            """Converte para booleano"""
            if not v:
                return False
            t = str(v).strip().lower()
            return t.startswith("s") or t == "1" or "yes" in t or "sim" in t

        # Loop principal
        self.stdout.write(f"Iniciando importação de {path}...")
        
        for row in range(2, ws.max_row + 1):
            try:
                # Ler número do pedido
                num_pedido = ws.cell(row=row, column=COL_PEDIDO).value
                if not num_pedido:
                    continue

                try:
                    num_pedido = int(float(str(num_pedido)))
                except:
                    self.stderr.write(f"Linha {row}: Número de pedido inválido")
                    continue

                # Ler todos os campos da linha
                data_pedido_raw = ws.cell(row=row, column=COL_DATA_PEDIDO).value
                solicitante_raw = ws.cell(row=row, column=COL_SOLICITANTE).value
                operador = ws.cell(row=row, column=COL_OPERADOR).value
                anv = ws.cell(row=row, column=COL_ANV).value
                tipo_atd_raw = ws.cell(row=row, column=COL_TIPO_ATD).value
                tipo_ped_raw = ws.cell(row=row, column=COL_TIPO_PED).value
                mpn = ws.cell(row=row, column=COL_MPN).value
                nome_item = ws.cell(row=row, column=COL_NOME).value
                qtd = ws.cell(row=row, column=COL_QTD).value
                doc_raw = ws.cell(row=row, column=COL_DOC_TS).value
                motivo = ws.cell(row=row, column=COL_MOTIVO).value
                obs = ws.cell(row=row, column=COL_OBS).value
                failure_description = ws.cell(row=row, column=COL_DESC_PANE).value
                troubleshooting = ws.cell(row=row, column=COL_TROUBLESHOOTING).value
                tsn = ws.cell(row=row, column=COL_TSN).value
                tso = ws.cell(row=row, column=COL_TSO).value
                sn = ws.cell(row=row, column=COL_SN).value
                anv_tsn = ws.cell(row=row, column=COL_ANV_TSN).value
                venc_raw = ws.cell(row=row, column=COL_VENC).value
                dest_anv = ws.cell(row=row, column=COL_DEST_ANV).value
                pn_alt1 = ws.cell(row=row, column=COL_PN_ALT1).value
                pn_alt2 = ws.cell(row=row, column=COL_PN_ALT2).value
                dpe = ws.cell(row=row, column=COL_DPE).value
                log = ws.cell(row=row, column=COL_LOG).value
                gmm = ws.cell(row=row, column=COL_GMM).value
                collected = ws.cell(row=row, column=COL_COLLECTED).value
                status_raw = ws.cell(row=row, column=COL_STATUS).value
                nf = ws.cell(row=row, column=COL_NF).value
                data_atd_raw = ws.cell(row=row, column=COL_DATA_ATD).value
                contrato_ant_raw = ws.cell(row=row, column=COL_CONTRATO_ANT).value
                note = ws.cell(row=row, column=COL_NOTES).value

                # Conversões de datas
                order_date = to_date(data_pedido_raw)
                if not order_date:
                    self.stderr.write(f"Linha {row}: Data de pedido inválida")
                    continue

                attended_date = to_date(data_atd_raw)
                venc_date = to_date(venc_raw)

                # Processamentos
                requester = parse_requester(solicitante_raw)
                order_type = parse_order_type(tipo_ped_raw)
                status = parse_status(status_raw)
                
                # Aeronaves
                anv_obj = aircraft_filter(anv)
                dest_anv_obj = aircraft_filter(dest_anv)

                # DOC e TEC_PUB
                doc_final, tec_pub_final = parse_doc_tec_pub(doc_raw)

                # Criar/atualizar ORDER
                order, created = Order.objects.get_or_create(
                    order_number=num_pedido,
                    order_year=order_date.year,
                    defaults={
                        "order_date": order_date,
                        "requester": requester,
                        "order_type": order_type,
                        "status": status,
                        "notes": clean_text(note) or "",
                        "created_by": user,
                        "updated_by": user,
                    },
                )

                if created:
                    created_orders += 1
                else:
                    # Atualizar campos se necessário
                    changed = False
                    if requester and not order.requester:
                        order.requester = requester
                        changed = True
                    if order_type and not order.order_type:
                        order.order_type = order_type
                        changed = True
                    if status != order.status:
                        order.status = status
                        changed = True
                    if note:
                        new_notes = clean_text(note)
                        if new_notes and new_notes not in (order.notes or ""):
                            order.notes = (order.notes or "") + "\n" + new_notes
                            changed = True
                    
                    if changed:
                        order.updated_by = user
                        order.save()
                        updated_orders += 1

                # Processar ITEM
                item_obj = None
                if mpn:
                    mpn_str = clean_text(mpn)
                    if mpn_str:
                        item_obj, item_created = Item.objects.get_or_create(
                            mpn=mpn_str,
                            defaults={
                                "name": clean_text(nome_item) or mpn_str,
                                "doc": doc_final,
                                "tec_pub": tec_pub_final,
                                "created_by": user,
                            },
                        )

                        # Atualizar doc/tec_pub se necessário
                        item_changed = False
                        if doc_final and item_obj.doc != doc_final:
                            item_obj.doc = doc_final
                            item_changed = True
                        if tec_pub_final and item_obj.tec_pub != tec_pub_final:
                            item_obj.tec_pub = tec_pub_final
                            item_changed = True
                        if item_changed:
                            item_obj.save()

                # Campos adicionais
                contract_old = parse_boolean(contrato_ant_raw)
                tsn_val = to_decimal(tsn)
                tso_val = to_decimal(tso)
                quantity = to_int(qtd) or 1
                

                # Criar ORDERITEM
                OrderItem.objects.create(
                    order=order,
                    item_item=item_obj,
                    operator=clean_text(operador),
                    aircraft=anv_obj,
                    aircraft_destination=dest_anv_obj,
                    service_type=clean_text(tipo_atd_raw),
                    quantity=quantity,
                    dpe=clean_text(dpe),
                    sn_attended = clean_text(sn),
                    nf_answer=clean_text(nf),
                    expiration_date_attended=venc_date,
                    attended_date=attended_date,
                    contract_old=contract_old,
                    reason=clean_text(motivo),
                    failure_description=clean_text(failure_description),
                    troubleshooting=clean_text(troubleshooting),
                    observation=clean_text(obs) or None,
                    created_by=user,
                    log = parse_boolean(log),
                    gmm=clean_text(gmm),
                    collected=parse_boolean(collected),
                    tsn_item=tsn_val,
                    tso_item=tso_val,
                    notes = note,
                )
                created_items += 1

                # Log de progresso a cada 100 linhas
                if row % 100 == 0:
                    self.stdout.write(f"Processadas {row} linhas...")

            except Exception as e:
                errors += 1
                self.stderr.write(
                    self.style.ERROR(f"Erro na linha {row}: {str(e)}")
                )

        # Relatório final
        self.stdout.write(
            self.style.SUCCESS(
                f"\n{'='*60}\n"
                f"IMPORTAÇÃO CONCLUÍDA\n"
                f"{'='*60}\n"
                f"📦 Pedidos criados: {created_orders}\n"
                f"🔄 Pedidos atualizados: {updated_orders}\n"
                f"📋 Itens criados: {created_items}\n"
                f"❌ Erros: {errors}\n"
                f"{'='*60}"
            )
        )